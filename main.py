import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkcalendar import DateEntry
import pandas as pd
from sqlalchemy import create_engine, text
from sqlalchemy.engine import URL
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image, ImageTk
from datetime import datetime
import os
import sys

# -------------------- APP INFO --------------------
APP_COMPANY = "Premium Industrial Solutions Pvt. Ltd."
APP_NAME = "Database to Excel Report Generator"
APP_VERSION = "v1.0.0"
APP_PROJECT_NAME = "BBPL KINLEY RO & MIS"
APP_DEVELOPER = "Developed by Subhajit Duttagupta"
APP_UPDATE = "UPDATED BY SEKH SAHABAJ"


# ================= PATH HANDLING =================
if getattr(sys, 'frozen', False):
    application_path = sys._MEIPASS
else:
    application_path = os.path.dirname(os.path.abspath(__file__))

LOGO_PATH = os.path.join(application_path, "logo.png")
TEMPLATE_PATH = os.path.join(application_path, "Report_Template.xlsx")

DEFAULT_REPORT_DIR = os.path.join(os.path.expanduser("~"), "Reports")
os.makedirs(DEFAULT_REPORT_DIR, exist_ok=True)

def get_writable_template_path():
    temp_dir = os.path.join(os.path.expanduser("~"), "Reports", "_template")
    os.makedirs(temp_dir, exist_ok=True)

    writable_template = os.path.join(temp_dir, "Report_Template.xlsx")

    if not os.path.exists(writable_template):
        import shutil
        shutil.copy(TEMPLATE_PATH, writable_template)

    return writable_template

# ✅ CONFIG FILE TO REMEMBER LAST SAVE PATH
CONFIG_PATH = os.path.join(os.path.expanduser("~"), ".report_app_path.txt")

# ================= DATABASE CONFIG =================
DATABASES = {
    "KINLEY_MIS_DB": {
        "name": "KINLEY_MIS_DB",
        "server": r"DESKTOP-87HT9VP\WINCC",
        "driver": "ODBC Driver 17 for SQL Server",
        "username": "admin",
        "password": "admin"
    },
    "KINLEY_RO_DB": {
        "name": "KINLEY_RO_DB",
        "server": r"DESKTOP-87HT9VP\WINCC",
        "driver": "ODBC Driver 17 for SQL Server",
        "username": "admin",
        "password": "admin"
    }
}

DATABASE_TABLES = {
    "KINLEY_MIS_DB": ["CP_CPK", "Mis_mean_stdev", "mis_tab"],
    "KINLEY_RO_DB": ["all_data", "RIO_DATA", "RO_CP_CPK", "RO_mean_stdev", "ro_tab", "TOTALIZER"]
}

# ================= MAIN APP =================
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Database Query to Excel")
        self.root.resizable(False, False)

        self.uploaded_image_path = None

        # ✅ LOAD LAST SAVE PATH
        if os.path.exists(CONFIG_PATH):
            with open(CONFIG_PATH, "r") as f:
                self.save_dir = f.read().strip() or DEFAULT_REPORT_DIR
        else:
            self.save_dir = DEFAULT_REPORT_DIR

        # ---------- LOGO ----------
        try:
            img = Image.open(LOGO_PATH).resize((200, 100))
            self.logo_img = ImageTk.PhotoImage(img)
            tk.Label(root, image=self.logo_img).grid(row=0, columnspan=4, pady=5)
            tk.Label(root, text="pisplindustry.com").grid(row=1, columnspan=4)
        except:
            pass

        # ---------- DATE / TIME ----------
        tk.Label(root, text="Start Date").grid(row=2, column=0)
        self.start_date = DateEntry(root, date_pattern="yyyy-mm-dd")
        self.start_date.grid(row=2, column=1)

        tk.Label(root, text="Start Time").grid(row=2, column=2)
        self.start_time = ttk.Combobox(root, values=[f"{h:02}:00:00" for h in range(24)])
        self.start_time.set("00:00:00")
        self.start_time.grid(row=2, column=3)

        tk.Label(root, text="End Date").grid(row=3, column=0)
        self.end_date = DateEntry(root, date_pattern="yyyy-mm-dd")
        self.end_date.grid(row=3, column=1)

        tk.Label(root, text="End Time").grid(row=3, column=2)
        self.end_time = ttk.Combobox(root, values=[f"{h:02}:00:00" for h in range(24)])
        self.end_time.set("23:59:59")
        self.end_time.grid(row=3, column=3)

        # ---------- DATABASE ----------
        tk.Label(root, text="Database").grid(row=4, column=0)
        self.database_var = tk.StringVar()
        self.database_box = ttk.Combobox(root, textvariable=self.database_var, values=list(DATABASES.keys()))
        self.database_box.set(list(DATABASES.keys())[0])
        self.database_box.grid(row=4, column=1)
        self.database_box.bind("<<ComboboxSelected>>", lambda e: self.populate_tables())

        tk.Label(root, text="Table").grid(row=5, column=0)
        self.table_var = tk.StringVar()
        self.table_box = ttk.Combobox(root, textvariable=self.table_var)
        self.table_box.grid(row=5, column=1)
        self.populate_tables()

        # ---------- IMAGE ----------
        tk.Label(root, text="Upload Image").grid(row=6, column=0)
        img_frame = tk.Frame(root)
        img_frame.grid(row=6, column=1, columnspan=2, sticky="w")
        tk.Button(img_frame, text="Browse", command=self.upload_image).pack(side=tk.LEFT)
        self.image_label = tk.Label(img_frame, text="No image selected", fg="gray")
        self.image_label.pack(side=tk.LEFT, padx=5)

        # ---------- SAVE PATH ----------
        tk.Label(root, text="Save Location").grid(row=7, column=0)
        path_frame = tk.Frame(root)
        path_frame.grid(row=7, column=1, columnspan=2, sticky="w")
        tk.Button(path_frame, text="Browse", command=self.browse_save_path).pack(side=tk.LEFT)
        self.path_label = tk.Label(path_frame, text=self.save_dir, fg="blue", wraplength=300)
        self.path_label.pack(side=tk.LEFT, padx=5)

        # ---------- BUTTONS ----------
        tk.Button(root, text="Generate Excel", command=self.generate_excel, width=20).grid(row=8, columnspan=4, pady=10)
        tk.Button(root, text="About", command=self.show_about).grid(row=9, columnspan=4)

    # ================= FUNCTIONS =================
    def populate_tables(self):
        tables = DATABASE_TABLES.get(self.database_var.get(), [])
        self.table_box["values"] = tables
        if tables:
            self.table_box.set(tables[0])

    def upload_image(self):
        file = filedialog.askopenfilename(
            filetypes=[("Image Files", "*.png *.jpg *.jpeg *.bmp")]
        )
        if file:
            self.uploaded_image_path = file
            self.image_label.config(text=os.path.basename(file), fg="green")

    def browse_save_path(self):
        path = filedialog.askdirectory(initialdir=self.save_dir)
        if path:
            self.save_dir = path
            self.path_label.config(text=path)

            # ✅ SAVE PATH FOR NEXT RUN
            with open(CONFIG_PATH, "w") as f:
                f.write(path)

    def add_image_to_excel(self, ws):
        if self.uploaded_image_path:
            img = ExcelImage(self.uploaded_image_path)
            img.width = 80
            img.height = 80
            ws.add_image(img, "O2")

    def generate_excel(self):
        start_dt = f"{self.start_date.get()} {self.start_time.get()}"
        end_dt = f"{self.end_date.get()} {self.end_time.get()}"
        table = self.table_var.get()
        db_name = self.database_var.get()

        if not table:
            messagebox.showerror("Error", "Table not selected")
            return

        if table in ["mis_tab", "all_data", "TOTALIZER", "RIO_DATA"]:
            date_style = 105
        else:
            date_style = 120

        start_sql = datetime.strptime(start_dt, "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S") if date_style == 105 else start_dt
        end_sql = datetime.strptime(end_dt, "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S") if date_style == 105 else end_dt

        db = DATABASES[db_name]

        conn_url = URL.create(
            "mssql+pyodbc",
            username=db["username"],
            password=db["password"],
            host=db["server"],
            database=db["name"],
            query={"driver": db["driver"], "TrustServerCertificate": "yes"}
        )

        engine = create_engine(conn_url)

        query = text(f"""
            SELECT *
            FROM [{db['name']}].dbo.[{table}]
            WHERE TRY_CONVERT(DATETIME, DateAndTime, {date_style}) >= TRY_CONVERT(DATETIME, :s, {date_style})
            AND TRY_CONVERT(DATETIME, DateAndTime, {date_style}) <= TRY_CONVERT(DATETIME, :e, {date_style})
            ORDER BY TRY_CONVERT(DATETIME, DateAndTime, {date_style}) DESC
        """)

        df = pd.read_sql_query(query, engine, params={"s": start_sql, "e": end_sql})

        wb = load_workbook(get_writable_template_path())
        ws = wb.active

        ws["A6"] = db_name
        ws["A6"].font = Font(bold=True, size=14)
        ws["A6"].alignment = Alignment(horizontal="center", vertical="center")

        ws["A7"] = table
        ws["A7"].font = Font(bold=True, size=12)
        ws["A7"].alignment = Alignment(horizontal="center", vertical="center")

        ws["A8"] = "Start DateTime:"
        ws["B8"] = start_dt
        ws["C8"] = "End DateTime:"
        ws["D8"] = end_dt

        for cell in ["A8", "B8", "C8", "D8"]:
            ws[cell].font = Font(bold=True)
            ws[cell].alignment = Alignment(horizontal="center", vertical="center")

        for c, col in enumerate(df.columns, 1):
            ws.cell(row=9, column=c, value=col).font = Font(bold=True)
            ws.cell(row=9, column=c).alignment = Alignment(horizontal="center")

        for r, row in enumerate(df.itertuples(index=False), 10):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
                ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

        self.add_image_to_excel(ws)
        
        print("Template path:", TEMPLATE_PATH)
        print("Save dir:", self.save_dir)
        
        filename = f"{table}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        out_path = os.path.join(self.save_dir, filename)
        wb.save(out_path)

        messagebox.showinfo("Success", f"Excel generated:\n{out_path}")

    # ---------------- ABOUT WINDOW ----------------
    def show_about(self):
        about = tk.Toplevel(self.root)
        about.title("About")
        about.resizable(False, False)
        about.transient(self.root)
        about.grab_set()

        w, h = 360, 300
        x = (about.winfo_screenwidth() // 2) - (w // 2)
        y = (about.winfo_screenheight() // 2) - (h // 2)
        about.geometry(f"{w}x{h}+{x}+{y}")

        frame = tk.Frame(about, padx=20, pady=15)
        frame.pack(fill="both", expand=True)

        try:
            img = Image.open(os.path.join(application_path, "logo.png"))
            img = img.resize((120, 55), Image.Resampling.LANCZOS)
            self.about_logo = ImageTk.PhotoImage(img)
            tk.Label(frame, image=self.about_logo).pack(pady=(0, 10))
        except:
            pass

        tk.Label(frame, text=APP_COMPANY, font=("Arial", 11, "bold")).pack()
        tk.Label(frame, text=APP_NAME, font=("Arial", 10)).pack(pady=2)
        tk.Label(frame, text=f"Version: {APP_VERSION}", font=("Arial", 9)).pack()
        tk.Label(frame, text=f"PROJECT: {APP_PROJECT_NAME}", font=("Arial", 9)).pack()
        tk.Label(frame, text=APP_DEVELOPER, font=("Arial", 9)).pack(pady=10)
        tk.Label(frame, text=APP_UPDATE, font=("Arial", 9)).pack()
        ttk.Separator(frame).pack(fill="x", pady=8)
        tk.Button(frame, text="Close", width=12, command=about.destroy).pack()


# ================= RUN =================
root = tk.Tk()
App(root)
root.mainloop()
